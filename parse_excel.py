import json
from openpyxl import load_workbook
from shifr import encrypt
def file_doc():
    # загружаем файл excel
    wb = load_workbook(filename = 'questions.xlsx')
    # выбираем рабочий лист
    worksheet = wb['Лист1']

    # так можно получить списком заголовки таблицы
    # headings = [c.value for c in next(worksheet.iter_rows(min_row=1, max_row=1))]
    table = []
    row=2
    max_row=worksheet.max_row #сколько строк в таблице
    temp={}
    # цикл по всем строкам таблицы
    while row<=max_row:
        # если значение ячейки с вопросом не пустое
        if worksheet.cell(row,1).value:
            # если временный словарь не пустой, то добавляем его в таблицу
            if temp: table.append(temp)
            temp={}
            # цикл по ячейкам строки от 1 до 4
            for i in range(1,5):
                # если ячейка варианта ответа или правильного ответа, то добавляем список
                if i in [3,4]:
                    temp[worksheet.cell(1, i).value]=[worksheet.cell(row, i).value]
                # иначе добавляем просто значение
                else:
                    temp[worksheet.cell(1, i).value]=worksheet.cell(row, i).value
        # если значение ячейки с вопросом пустое, значит идёт перечисление вариантов ответа или ответов
        else:
            # если значение ячейки непустое, то добавляем в список
            if worksheet.cell(row, 3).value: temp[worksheet.cell(1, 3).value].append(worksheet.cell(row, 3).value)
            if worksheet.cell(row, 4).value: temp[worksheet.cell(1, 4).value].append(worksheet.cell(row, 4).value)
        row+=1
    else:
        # добавляем в список последний набранный временный словарь
        table.append(temp)   

    with open('questions_from_excel.json','w', encoding='utf-8') as f:
        json.dump(table,f, ensure_ascii=False, indent=4)

    # with open('key.key', 'rb') as f:
    #     key= f.read()
    # encrypt('questions_from_excel.json', key)

file_doc()